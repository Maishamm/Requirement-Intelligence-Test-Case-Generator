from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any


def export_test_cases_json(output: Any, file_path: str | Path) -> Path:
    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(output, indent=2, ensure_ascii=False), encoding="utf-8")
    return path


def export_test_cases_excel(output: Any, file_path: str | Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")

    feature_font = Font(name="Segoe UI", size=11, bold=True, color="1F497D")
    feature_fill = PatternFill(start_color="E9EEF4", end_color="E9EEF4", fill_type="solid")

    data_font = Font(name="Segoe UI", size=10)
    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    headers = ["TC ID", "Title", "Type", "Source Requirement", "Evidence Quote", "Preconditions", "Steps", "Expected Result"]
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    features_list = []
    if isinstance(output, dict):
        if "features" in output:
            features_list = output["features"]
        else:
            features_list = [output]
    else:
        features_list = [{"feature_id": "Unknown", "feature_name": "Test Cases List", "test_cases": output}]

    row_idx = 2
    zebra_counter = 0
    feature_rows = set()

    for feat in features_list:
        if not isinstance(feat, dict):
            continue

        feature_id = feat.get("feature_id", "")
        feature_name = feat.get("feature_name", "")
        test_cases = feat.get("test_cases", [])

        if not test_cases:
            continue

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        feature_rows.add(row_idx)
        feat_cell = ws.cell(row=row_idx, column=1)
        feat_cell.value = f"Feature {feature_id}: {feature_name}"
        feat_cell.font = feature_font
        feat_cell.fill = feature_fill
        feat_cell.alignment = Alignment(vertical="center", indent=1)
        feat_cell.border = thin_border

        for col_idx in range(2, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = thin_border
            ws.cell(row=row_idx, column=col_idx).fill = feature_fill

        ws.row_dimensions[row_idx].height = 24
        row_idx += 1

        for tc in test_cases:
            if not isinstance(tc, dict):
                continue

            tc_id = tc.get("test_case_id", f"TC-{zebra_counter+1:03d}")
            title = tc.get("title", "")
            tc_type = tc.get("type", "")

            source_req = tc.get("source_requirement_ids", [])
            if isinstance(source_req, list):
                source_req_str = ", ".join(str(x) for x in source_req)
            else:
                source_req_str = str(source_req or "")

            evidence_quote = str(tc.get("evidence_quote", "") or "")

            preconditions = tc.get("preconditions", [])
            if isinstance(preconditions, list):
                preconditions_str = "\n".join(f"- {p}" for p in preconditions)
            else:
                preconditions_str = str(preconditions or "")

            steps = tc.get("steps", [])
            if isinstance(steps, list):
                steps_str = ""
                for s_idx, step in enumerate(steps, 1):
                    step_stripped = str(step).strip()
                    if re.match(r"^\d+\.", step_stripped):
                        steps_str += f"{step_stripped}\n"
                    else:
                        steps_str += f"{s_idx}. {step_stripped}\n"
                steps_str = steps_str.strip()
            else:
                steps_str = str(steps or "")

            expected = tc.get("expected_result", [])
            if isinstance(expected, list):
                expected_str = "\n".join(f"- {e}" for e in expected)
            else:
                expected_str = str(expected or "")

            row_data = [tc_id, title, tc_type, source_req_str, evidence_quote, preconditions_str, steps_str, expected_str]
            ws.append(row_data)

            row_fill = zebra_fill if zebra_counter % 2 == 1 else None

            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                if row_fill:
                    cell.fill = row_fill
                cell.border = thin_border

                if col_idx in [1, 3, 4]:
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            max_lines = max(
                len(preconditions_str.splitlines()),
                len(steps_str.splitlines()),
                len(expected_str.splitlines()),
                1
            )
            ws.row_dimensions[row_idx].height = max(18, 15 + max_lines * 13)

            row_idx += 1
            zebra_counter += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in feature_rows:
                continue
            val_str = str(cell.value or "")
            lines = val_str.splitlines()
            if lines:
                longest_line = max(len(l) for l in lines)
                if longest_line > max_len:
                    max_len = longest_line
            else:
                if len(val_str) > max_len:
                    max_len = len(val_str)

        if col_letter == "A":
            ws.column_dimensions[col_letter].width = 12
        elif col_letter == "C":
            ws.column_dimensions[col_letter].width = 15
        elif col_letter in ["B", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[col_letter].width = max(20, min(max_len + 3, 45))
        else:
            ws.column_dimensions[col_letter].width = max(10, max_len + 3)

    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path
